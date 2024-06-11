import csv
from io import BytesIO
from io import TextIOWrapper
from zipfile import ZipFile

import xlrd
from openpyxl import load_workbook
from pypdf import PdfReader


def test_open_csv_file_from_zip():
    with ZipFile('resources/my_archive.zip') as zip_file:
        with zip_file.open('tmp/file_example_CSV.csv') as csv_file:
            csvreader = list(csv.reader(TextIOWrapper(csv_file, 'utf-8-sig')))
            temp_row = csvreader[3]
            assert temp_row[1] == 'Достоевский Фёдор Михайлович'


def test_open_xlsx_file_from_zip():
    with ZipFile("resources/my_archive.zip") as zip_file:
        with zip_file.open('tmp/file_example_XLSX_50.xlsx') as xlsx_file:
            wb = load_workbook(xlsx_file)
            ws = wb.active
            assert ws["B7"].value == 'Gaston'
            assert ws["G27"].value == '16/08/2016'


def test_open_xls_file_from_zip():
    with ZipFile("resources/my_archive.zip") as zip_file:
        with zip_file.open('tmp/file_example_XLS_10.xls') as xls_file:
            in_memory_file = BytesIO(xls_file.read())
            workbook = xlrd.open_workbook(file_contents=in_memory_file.getvalue())
            sheet = workbook.sheet_by_index(0)
            assert sheet.cell(rowx=9, colx=1).value == 'Vincenza'
            assert sheet.cell(rowx=2, colx=4).value == 'Great Britain'


def test_open_pdf_file_from_zip():
    with ZipFile("resources/my_archive.zip") as zip_file:
        with zip_file.open("tmp/Python Testing with Pytest.pdf") as pdf_file:
            reader = PdfReader(pdf_file)
            number_of_pages = len(reader.pages)

            for page in reader.pages[42:43]:
                text_block = page.extract_text()

            assert number_of_pages == 256
            assert ('Testing a Package\n'
                    'We’ll use the sample project, Tasks, as discussed in \u200b The Tasks '
                    'Project \u200b, to see how to write test\n'
                    'functions for a Python package. Tasks is a Python package that includes a '
                    'command-line tool of\n'
                    'the same name, tasks.\n') in text_block
